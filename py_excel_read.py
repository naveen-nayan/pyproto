__author__ = 'naveen'
__date__ = '15/02/2020'


''' openpyxl to read and write data
    pip install openpyxl'''

# import pandas as pd

# excel_file = 'movies.xls'
# movie = pd.read_excel(excel_file)
# print(movie.head(0))


# import package
import openpyxl

# load workbook
workbook_path = "movies.xlsx"
# load sheet as workbook
sheet = openpyxl.load_workbook(workbook_path)
# print no of sheet
print(sheet.sheetnames)
# see active sheet
print(sheet.active.title)

# create objet of any sheet on which we want to perform action

sheet1 = sheet['1900s']
print(sheet1.title)

# sheet level object
print(sheet1['A3'].value)
print(sheet1['B3'].value)

# shell level approach
c1 = sheet1.cell(1,1)

# as keyword argument
C1 = sheet1.cell(column=1, row=1)
print(c1.value)

# find rows in sheet
rows  = sheet1.max_row
print(rows)
# find columns in sheet
columns = sheet1.max_column
print(columns)


# print all data in excel
for i in range(1,rows+1):
    for j in range(1,columns+1):
        print(sheet1.cell(column=j, row=i).value+" ", end='')
    print()


# Method2
for r in sheet1['A1': 'Y1339']:
    for c in r:
        print(c.value+ " ", end='')
    print()