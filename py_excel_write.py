__author__ = 'naveen'
__date__ = '15/02/2020'

import openpyxl

wk = openpyxl.Workbook()
sh1 = wk.active

# create sheet1
sh1.title = "Sheet1"

count = 1
for r in range(1, 50):
    for c in range(1, 5):
        sh1.cell(row=r, column=c).value = count
        count +=1

wk.create_sheet(title="Sheet2")
sh2 = wk['Sheet2']

for r in range(1, 50):
    for c in range(1, 5):
        sh2.cell(row=r, column=c).value = count
        count +=1


wk.save('textSheet.xlsx')
